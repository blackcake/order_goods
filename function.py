# -*- coding: utf-8 -*-
import json
import openpyxl
import os
import pandas
import re
import requests
import time
from openpyxl.utils.dataframe import dataframe_to_rows


def read_xlsx(path, sheet=None):
    # 读excel的通用函数
    workbook = openpyxl.load_workbook(path)
    if sheet:
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active
    data = worksheet.values
    cols = next(data)
    dataframe = pandas.DataFrame(data, columns=cols)
    return dataframe


def export_xlsx(path, dataframe):
    # 写入excel的通用函数
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        sheet.append(row)
    if not os.path.exists(os.path.split(path)[0]):
        os.mkdir(os.path.split(path)[0])
    workbook.save(path)


class Comments:
    # 抓取QQ群相册的评论，如果目标文件夹下没有相册评论.xlsx，生成并写入记录，如果存在，对比记录，并在最后一条记录后添加新增评论
    # 如果上一次执行时的最后一条评论，在第二次执行时已经被删除，第二次执行会在最后一条记录后新增所有评论，（其实加一行去重就可以解决）
    def __init__(self, album, login, path_project, topicid):
        self._album = album
        self._album['topicId'] = topicid
        self._login = login
        self._path_project = path_project
        self.path_comments = os.path.join(path_project, '相册评论.xlsx')
        self.count_comment = 0
        self.count_history = 0
        self.count_new = 0

    def get_comments_history(self):
        # 读取相册评论.xlsx，如果不存在，返回空dataframe
        if not os.path.exists(self.path_comments):
            return pandas.DataFrame([])
        dataframe = read_xlsx(self.path_comments)
        self.count_history = len(dataframe)
        return dataframe

    def _get_data(self, start):
        print('正在抓取第%d页的评论' % int(start/10+1))
        # 爬虫，获取某一页的相册评论
        url = 'https://h5.qzone.qq.com/proxy/domain/u.photo.qzone.qq.com/cgi-bin/upp/qun_list_photocmt_v2'
        query_string = {
            'uin': self._album['uin'],
            'hostUin': self._album['hostUin'],
            'start': '{:d}'.format(start),
            'num': '10',
            'order': '0',
            'topicId': self._album['topicId'],
            'format': 'jsonp',
            'inCharset': 'utf-8',
            'outCharset': 'utf-8',
            'ref': 'qunphoto',
            'random': self._login['random'],
            'g_tk': self._login['g_tk'],
            'qzonetoken': self._login['qzonetoken']
        }
        header = {
            'cookie': self._login['cookie']
        }
        data = requests.request('GET', url, headers=header, params=query_string).text
        # 返回结果外面有个_Callback()
        return json.loads(data[10: -2])['data']

    def get_comments_now(self):
        # 爬虫，抓取当前相册所有评论，输出dataframe
        data = self._get_data(0)
        comments = data['comments']
        for i in range(10, data['total'], 10):
            # time.sleep(1)
            comments.extend(self._get_data(i)['comments'])
        dict_comments = {
            'content': [c['content'] for c in comments],
            'id': [c['id'] for c in comments],
            'postTime': [c['postTime'] for c in comments],
            'name': [c['poster']['name'] for c in comments],
        }
        dataframe = pandas.DataFrame(dict_comments)[['name', 'id', 'postTime', 'content']]
        dataframe['postTime'] = dataframe['postTime'].apply(lambda x: time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(x)))
        dataframe.insert(len(dataframe.columns), '相册评论（可修改）', dataframe['content'])
        self.count_comment = len(dataframe)
        return dataframe.rename(columns={'name': '群昵称', 'postTime': '评论时间', 'content': '相册评论（请勿修改）'})

    def get_comments_concat(self):
        # 合并已有的excel和抓取到的评论，在最后一行的基础上添加新增评论
        df_history = self.get_comments_history()
        df_now = self.get_comments_now()
        if df_history.empty:
            self.count_new = len(df_now)
            return df_now
        row = df_history[['群昵称', 'id', '评论时间', '相册评论（请勿修改）']].loc[df_history.index[-1]].tolist()
        i = -1
        for i in df_now.index:
            if df_now[['群昵称', 'id', '评论时间', '相册评论（请勿修改）']].loc[i].tolist() == row:
                break
        if df_now[['群昵称', 'id', '评论时间', '相册评论（请勿修改）']].loc[i].tolist() == row:
            self.count_new = len(df_now) - i - 1
            return df_history.append(df_now[i+1:])[df_history.columns]
        dataframe = df_history.append(df_now)[df_history.columns]
        dataframe.drop_duplicates(subset=['群昵称', 'id', '评论时间', '相册评论（请勿修改）'], inplace=True)
        self.count_new = len(dataframe) - len(df_history)
        return dataframe

    def export(self):
        # 合并新评论和之前的记录后，生成新的excel，覆盖原有文件
        dataframe = self.get_comments_concat()
        export_xlsx(self.path_comments, dataframe)


class NormalizedString:
    def __init__(self, dataframe_goods, goods_type, quantity_max):
        self._dataframe_goods = dataframe_goods
        self._goods_type = goods_type
        self._quantity_max = quantity_max

    def _match_goods_role(self, row, number):
        string = row[0]
        list_goods = []
        for i in self._dataframe_goods.index:
            if self._dataframe_goods.loc[i, '别名'] in string and self._dataframe_goods.loc[i, '全称'] not in list_goods:
                list_goods.append(self._dataframe_goods.loc[i, '全称'])
                string = string.replace(self._dataframe_goods.loc[i, '别名'], '')
            if not string:
                break
        if not list_goods:
            return None
        if self._goods_type:
            result = [','.join([goods, row[1], number]) for goods in list_goods]
        else:
            result = [','.join([goods, number]) for goods in list_goods]
        return ';\n'.join(result)

    def _match_goods_type(self, string):
        # 类型解析，如果不区分柄，直接返回[[角色]]；如果区分柄类型的话，以类型作为标识分割字符串，返回格式[[角色, 类型]]，如果所有类型都要，类型标为'all'
        # 如果结果是[柄1][角色1][柄2]，拆分出来的柄多于角色，会直接报错；[角色1][柄1][角色2]，[角色2]会记为all
        if not self._goods_type:
            return [[string]]
        list_role = [row for row in re.split('[%s]+' % ''.join(set(list(''.join(self._goods_type)))),
                                             string.strip()) if row]
        list_type = [row for row in re.split('[^%s]+' % ''.join(set(list(''.join(self._goods_type)))),
                                             string.strip()) if row]
        if len(list_role) < len(list_type):
            return 'Error'
        list_role_norm = []
        for i in range(len(list_role)):
            if i >= len(list_type):
                list_role_norm.append([list_role[i], 'all'])
                continue
            role_type = [t for t in self._goods_type if t in list_type[i]]
            if len(role_type) == len(self._goods_type):
                list_role_norm.append([list_role[i], 'all'])
            else:
                list_role_norm.extend([[list_role[i], t] for t in role_type])
        return list_role_norm

    def _match_goods(self, string, number):
        # 读取角色清单，根据别名对实际名称进行匹配（只用别名列进行匹配，所以至少需要一个包含全称任何一个字的别名）
        # 如果排的是一个队伍，为了方便核对，如vk1，只会转换成valkyrie1，并不会拆分成斋宫宗1和影片美伽1
        list_role_norm = self._match_goods_type(string)
        if list_role_norm == 'Error':
            return 'Error'
        result = []
        for row in list_role_norm:
            result.append(self._match_goods_role(row, number))
        return ';\n'.join([role for role in result if role])

    def _match_number(self, string):
        list_goods = [row for row in re.split('[0-9凹|余]+', string) if row]
        list_number = [row for row in re.split('[^0-9凹|余]+', string) if row]
        # 把凹余/余凹/余/凹统一成一种表述
        list_number = [str(self._quantity_max) if '余' in row or '凹' in row else row for row in list_number]
        return list_goods, list_number

    def _normalize_string_buy(self, string):
        # 规范化用户点单要求的字符串，如vk和kn各1，改成valkyrie1+knights1
        # 先不考虑转单的
        # 暂时只考虑数量后置的
        # all有3个字符，替换
        string = string.strip().lower().replace('all', '凹').replace('.', '').replace('。', '')
        # 如果角色和数字对不起来或者识别到的角色数量为0，返回错误，手动填
        list_goods, list_number = self._match_number(string)
        if len(list_goods) < len(list_number) or not list_goods:
            return 'Error'

        result_match = [self._match_goods(list_goods[i], list_number[i]) for i in range(len(list_number))]
        order = ';\n'.join([r for r in result_match if r])
        if 'Error' in order or not order:
            return 'Error'
        return order

    def _normalize_string_undo(self, string):
        result = self._normalize_string_buy(string)
        if result == 'Error':
            return 'Error'
        newresult = []
        for row in result.split(';\n'):
            list_row = row.split(',')
            list_row[-1] = str(int(list_row[-1]) * -1)
            newresult.append(','.join(list_row))
        return ';\n'.join(newresult)

    def normalize_goods(self, string):
        # 规范化用户点单要求的字符串，如vk和kn各1，改成valkyrie1+knights1
        # 先不考虑转单的
        # 暂时只考虑数量后置的
        if string in ('Error', 'Done'):
            return string
        if string.startswith('撤'):
            return self._normalize_string_undo(string)
        return self._normalize_string_buy(string)


class OrderRecord:
    # 在爬虫抓取的相册评论的基础上，把相册评论转换成规范化的点单记录，把评论用户的ID分割成正常的名字
    # 在相册评论.xlsx中新增对应列[圈名]和[评论规范化]；如果相册评论.xlsx不存在，直接报错；
    # 如果已有对应列，原有记录不会修改，只会对[评论规范化]这一列原本为空的行做修改
    def __init__(self, myname, path_project, path_goods, goods_type, quantity_max):
        self._myname = myname
        self.path_comments = os.path.join(path_project, '相册评论.xlsx')
        self._df_goods = read_xlsx(path_goods, '名称')
        self._goods_type = goods_type
        self._quantity_max = quantity_max
        self.dataframe_error = pandas.DataFrame([])
        self.count_comment = 0

    def get_comments(self):
        if not os.path.exists(self.path_comments):
            raise ValueError('找不到文件')
        dataframe = read_xlsx(self.path_comments)
        if '圈名' not in dataframe.columns:
            dataframe.insert(len(dataframe.columns), '圈名', None)
        if '评论规范化' not in dataframe.columns:
            dataframe.insert(len(dataframe.columns), '评论规范化', None)
        if '转单人' not in dataframe.columns:
            dataframe.insert(len(dataframe.columns), '转单人', '')
        self.count_comment = len(dataframe)
        return dataframe

    def _normalize_user(self, string):
        # 规范化用户名称，把自己的id替换成圈名
        if string == self._myname['name']:
            return self._myname['nickname']
        return re.split('[- （(@【—_–，。!]', string)[0]

    def _get_transfer(self, dataframe):
        dataframe.reset_index(drop=True)
        # 转单
        for i in dataframe.index:
            string = dataframe.loc[i, '相册评论（可修改）']
            if not string.startswith('转'):
                continue
            if '给' not in string:
                dataframe.loc[i, '评论规范化'] = 'Error'
            # 支持格式"转[谷]给[转单人]"
            username = dataframe.loc[i, '圈名']
            transfer = string.split('给')[-1]
            flag = 0
            for j in dataframe[i + 1:].index:
                if transfer.startswith(dataframe.loc[j, '圈名']) and dataframe.loc[j, '评论规范化'] != 'Done' \
                        and dataframe.loc[j, '相册评论（可修改）'].startswith('接' + username):
                    flag = 1
                    transfer = dataframe.loc[j, '圈名']
                    dataframe.loc[j, '评论规范化'] = 'Done'
                    break
            if flag:
                dataframe.loc[i, '转单人'] = transfer
            else:
                dataframe.loc[i, '评论规范化'] = 'Error'
        # 以防把没匹配到的接转单识别成排谷
        for i in dataframe.index:
            if dataframe.loc[i, '相册评论（可修改）'].startswith('接') and dataframe.loc[i, '评论规范化'] != 'Done':
                dataframe.loc[i, '评论规范化'] = 'Error'
        return dataframe

    def get_comments_normalize(self):
        normalized = NormalizedString(self._df_goods, self._goods_type, self._quantity_max)
        # 读取相册评论.xlsx，根据[相册评论（可修改）]的值对字符串进行规范化处理，只对[评论规范化]列为空的数据进行
        dataframe = self.get_comments()
        dataframe['圈名'] = dataframe['群昵称'].apply(self._normalize_user)
        dataframe['评论规范化'] = dataframe['相册评论（可修改）'].apply(normalized.normalize_goods)
        # 转接
        dataframe = self._get_transfer(dataframe)
        # 报错
        self.dataframe_error = dataframe[dataframe['评论规范化'] == 'Error']
        return dataframe

    def export(self):
        # 合并新处理的评论和之前的记录后，生成新的excel，覆盖原有文件
        dataframe = self.get_comments_normalize()
        export_xlsx(self.path_comments, dataframe)


class OrderList:
    # 读取相册评论.xlsx，匹配队伍中的所有角色，生成排表和每个人的总数价格，这一步永远是全量更新
    # TODO：生成余量图
    def __init__(self, path_goods, path_project, quantity_min, quantity_max, goods_type):
        self._path_project = path_project
        self._path_comments = os.path.join(path_project, '相册评论.xlsx')
        self._df_team = read_xlsx(path_goods, '队伍')
        self._path_template = os.path.join(path_project, '模板.xlsx')
        self._quantity_min = quantity_min
        self._quantity_max = quantity_max
        self._goods_type = goods_type

    def get_comments(self):
        # 读取相册评论.xlsx，输出dataframe
        if not os.path.exists(self._path_comments):
            raise ValueError('找不到文件')
        dataframe = read_xlsx(self._path_comments)[['圈名', '评论规范化', '转单人']].rename(
            columns={'圈名': 'name', '评论规范化': 'order', '转单人': 'transfer'})
        dataframe['transfer'] = dataframe['transfer'].fillna('')
        return dataframe[~dataframe['order'].isin(['Error', 'Done'])].reset_index(drop=True)

    def get_comments_detail(self):
        # 把规范化评论里的结果，拆分成角色、类型（如果不区分就没有）和数量
        dataframe = self.get_comments()
        # 点单记录, 一行拆分成多行
        line_split = dataframe['order'].str.split(';\n', expand=True).stack().reset_index(level=1, drop=True)
        dataframe = dataframe[['name', 'transfer']].join(line_split.rename('order'))
        if self._goods_type:
            dataframe[['role', 'type', 'quantity']] = dataframe['order'].str.split(',', 2, expand=True)
        else:
            dataframe[['role', 'quantity']] = dataframe['order'].str.split(',', 1, expand=True)
        dataframe['quantity'] = dataframe['quantity'].apply(int)
        return dataframe.reset_index(drop=False)

    def _match_team(self, dataframe):
        # 对排了一整队的评论，匹配队伍中的所有人
        dataframe_match = pandas.merge(dataframe, self._df_team, left_on='role', right_on='队伍', how='left')
        for i in dataframe_match.index:
            if not pandas.isnull(dataframe_match.loc[i, '角色']):
                dataframe_match.loc[i, 'role'] = dataframe_match.loc[i, '角色']
        dataframe_match = dataframe_match.drop(['队伍', '角色'], axis=1)
        return dataframe_match

    def _match_type(self, dataframe):
        # 如果区分柄，根据模板中的角色和柄类型，匹配所有角色和柄
        # 主要是把all转换成对应的类型，但是如果goods_type参数和模板对不上，错的部分不会保留
        if not self._goods_type:
            return dataframe
        df_template = read_xlsx(self._path_template, '排表')
        dataframe_goods_all = pandas.merge(dataframe[dataframe['type'] == 'all'], df_template, left_on='role',
                                           right_on='角色', how='inner')[['index', 'name', 'role', '柄类型', 'quantity',
                                                                        'transfer']].rename(columns={'柄类型': 'type'})
        dataframe_goods_single = dataframe[dataframe['type'] != 'all'][['index', 'name', 'role', 'type', 'quantity',
                                                                        'transfer']]
        dataframe = pandas.concat([dataframe_goods_all, dataframe_goods_single])
        return dataframe.sort_values('index').drop('index', axis=1).reset_index(drop=True)

    def _namelist_buy(self, row, namelist):
        # 加购
        name, number = row.split(',')[0], int(row.split(',')[1])
        namelist.extend([name] * int(number / self._quantity_min))
        return namelist

    def _namelist_undo(self, row, namelist):
        # 撤排
        name, number = row.split(',')[0], int(row.split(',')[1])
        top, last = namelist[:], []
        if len(namelist) > self._quantity_max:
            top = namelist[:self._quantity_max]
            last = [i for i in namelist[self._quantity_max:] if i != name]
        top.reverse()
        for j in range(abs(number)):
            try:
                top.remove(name)
            except:
                print('警告: %s实际排到的数量小于撤排数' % name)
                break
        top.reverse()
        return top[:] + last[:]

    def _namelist_transfer(self, row, namelist):
        # 转单
        name, number, username = row.split(',')
        number = int(number)
        top, last = namelist[:], []
        if len(namelist) > self._quantity_max:
            top = namelist[:self._quantity_max]
            last = [i for i in namelist[self._quantity_max:] if i != name]
        top.reverse()
        for j in range(abs(number)):
            try:
                top[top.index(name)] = username
            except:
                print('警告: %s实际排到的数量小于转单数' % name)
                break
        top.reverse()
        return top[:] + last[:]

    def _generate_namelist(self, string):
        # 生成排表的名单
        namelist = []
        for row in string.split(';'):
            if row.split(',')[-1]:
                namelist = self._namelist_transfer(row, namelist)
            elif int(row.split(',')[1]) >= 0:
                namelist = self._namelist_buy(row, namelist)
            else:
                namelist = self._namelist_undo(row, namelist)
        if len(namelist) > self._quantity_max:
            namelist = namelist[:self._quantity_max]
        return ','.join(namelist)

    def _calculate_real_number(self, dataframe):
        # 根据设定的最大值计算每个人实际上排到的数量
        if self._goods_type:
            column_groupby = ['role', 'type']
        else:
            column_groupby = ['role']
        dataframe = dataframe.copy()
        dataframe.insert(0, 'namelist', dataframe.apply(
            lambda x: '%s,%s,%s' % (x['name'], int(x['quantity']/self._quantity_min), x['transfer']), axis=1))
        dataframe = dataframe.groupby(column_groupby).agg({'namelist': lambda x: ';'.join(x)}
                                                          ).reset_index(drop=False)
        dataframe['namelist'] = dataframe['namelist'].apply(self._generate_namelist)
        dataframe.insert(0, 'quantity', dataframe['namelist'].apply(lambda x: len(x.split(',')) if x else 0))
        return dataframe

    def _match_price(self, dataframe):
        # 计算每个人每类商品的数量和总价
        df_template = read_xlsx(self._path_template, '排表')
        if self._goods_type:
            column_left, column_right = ['role', 'type'], ['角色', '柄类型']
        else:
            column_left, column_right = ['role'], ['角色']
        # 把排表拆分成多行, 一个用户一行
        dataframe = dataframe[dataframe['quantity'] > 0]
        line_split = dataframe['namelist'].str.split(',', expand=True).stack().reset_index(level=1, drop=True)
        dataframe = dataframe[column_left].join(line_split.rename('name'))
        dataframe.insert(0, 'quantity', 1)
        dataframe = dataframe.groupby(['name']+column_left).agg({'quantity': 'sum'}).reset_index(drop=False)
        dataframe = pandas.merge(dataframe, df_template, left_on=column_left, right_on=column_right, how='inner')
        dataframe.insert(0, '总价', dataframe['单价'] * dataframe['quantity'])
        dataframe = dataframe[['name']+column_right+['quantity', '单价', '总价']].rename(
            columns={'name': '圈名', 'quantity': '数量'})
        dataframe = dataframe.sort_values(['圈名'] + column_right)
        return dataframe

    def get_comments_calc(self):
        # 整合流程，读取相册评论.xlsx，匹配队伍中的所有角色，匹配同一角色的不同柄，计算每个人实际排的数量，计算总价
        dataframe = self.get_comments_detail()
        dataframe = self._match_team(dataframe)
        dataframe = self._match_type(dataframe)
        dataframe_order = self._calculate_real_number(dataframe)
        dataframe_price = self._match_price(dataframe_order)
        return dataframe_order, dataframe_price

    def _match_index(self, dataframe):
        # 获取这个角色在模板.xlsx中实际的行数
        df_template = read_xlsx(self._path_template, '排表').reset_index(drop=False)
        if self._goods_type:
            column_left, column_right = ['role', 'type'], ['角色', '柄类型']
        else:
            column_left, column_right = ['role'], ['角色']
        dataframe = pandas.merge(dataframe, df_template, left_on=column_left, right_on=column_right, how='inner')
        dataframe.index = dataframe['index'] + 2
        return dataframe

    def export(self):
        # 输出结果
        dataframe_order, dataframe_price = self.get_comments_calc()
        dataframe_order = self._match_index(dataframe_order)
        workbook = openpyxl.load_workbook(self._path_template)
        # 排表
        sheet_role = workbook['排表']
        columns = next(sheet_role.values)
        index_role, index_sum, index_number = columns.index('角色')+1, columns.index('余量')+1, columns.index('余量')+2
        for i in range(2, sheet_role.max_row+1):
            if i in dataframe_order.index:
                sheet_role.cell(row=i, column=index_sum).value = self._quantity_max - dataframe_order.loc[i, 'quantity']
                namelist = dataframe_order.loc[i, 'namelist'].split(',')
                for j in range(len(namelist)):
                    sheet_role.cell(row=i, column=j+index_number).value = namelist[j]
            elif sheet_role.cell(row=i, column=index_role).value:
                sheet_role.cell(row=i, column=index_sum).value = self._quantity_max
        # 每个人每种商品的总数和总价
        sheet_person = workbook.create_sheet('个人')
        for row in dataframe_to_rows(dataframe_price, index=False, header=True):
            sheet_person.append(row)
        # 删除参数sheet
        sheet_params = workbook['参数']
        workbook.remove(sheet_params)
        workbook.save(os.path.join(self._path_project, '排表.xlsx'))
